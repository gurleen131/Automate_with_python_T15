import openpyxl

# Load the Excel file
workbook = openpyxl.load_workbook('Book1.xlsx')

# Select a specific worksheet
worksheet = workbook['Book1']

# Read data from a specific cell
value = worksheet.cell(row=1, column=1).value
value = worksheet.cell(row=2, column=2).value



# Print the value
print(value)
